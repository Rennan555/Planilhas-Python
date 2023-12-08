from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

wb = load_workbook("data/pivot_table.xlsx")
sheet = wb["Output"]

## Referências das linhas e colunas
min_col = wb.active.min_column
max_col = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

## Adicionando dados e cateforias
bc = BarChart()

data = Reference(
    sheet,
    min_col = min_col + 1, ## Pular a linha dos títulos
    max_col = max_col,
    min_row = min_row,
    max_row = max_row
)

categories = Reference(
    sheet,
    min_col = min_col,
    max_col = min_col,
    min_row = min_row + 1, ## Pular o cabeçalho
    max_row = max_row
)

bc.add_data(data, titles_from_data=True)
bc.set_categories(categories)

## Criando o gráfico
sheet.add_chart(bc, "A6")
bc.title = "GRAPHIC"
bc.style = 2

## Salvando o Work Book
wb.save("data/barchart.xlsx")
