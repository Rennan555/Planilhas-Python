from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("data/barchart.xlsx")
sheet = wb["Output"]

min_col = wb.active.min_column
max_col = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

## Incluindo fórmula
sheet["A5"] = "Total"
#sheet["B5"] = "=SUM(B2:B4)" -> Maneira rabusta

for i in range(min_col+1, max_col+1): ## Pula cabeçalho
    letter = get_column_letter(i)
    sheet["{0}{1}".format(letter, max_row+1)] = "=SUM({0}{1}:{0}{2})".format(letter, min_row, max_row)

wb.save("data/mod_barchart.xlsx")
